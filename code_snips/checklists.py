import os
import os.path
import sys
import json
import warnings
import argparse
import logging
from openpyxl import Workbook, load_workbook
from slugify import slugify
from openpyxl.utils import get_column_letter
# from openpyxl.utils.exceptions import InvalidFileException

checklist = "/Users/juliaguimiot/practicespace/p/Test_Checklist.xlsx"

def validate_accessories(accessories_str, sheet):
    valid_accessories=['display_text', 'notes', 'camera', 'gallery']
    accessories = accessories_str.split(',')
    final_accessories=[]

    for ac in accessories:
        try:
            if ac.strip().lower() in valid_accessories:
                final_accessories.append(ac.strip())
        except KeyError as e:
            logging.error("ERROR: {} is not a valid 'default accessories' entry. Please enter a valid list of accessories.".format(ac.strip(), e))
            checklist_errors.append("Invalid accessory: {} is not a valid 'default accessories' entry. Please enter a valid list of accessories.".format(ac.strip(), e))
    return final_accessories

def validate_line_item_type(line_item_type_str, sheet):
    tmpstr=""
    try:
        if line_item_type_str:
            tmpstr = "".join([x if ord(x) < 128 else '?' for x in line_item_type_str])
            tmpstr = tmpstr.strip().lower()
            valid_types=["text", "choice", "text date", "text block", "image", "multiple choice", "signature"]
            if tmpstr in valid_types:
                error=False
    except KeyError as e:
        logging.error("ERROR: Processing '{}' worksheet. {} is not a valid line item type. Valid line item types are: {}".format(sheet, tmpstr, valid_types))
        checklist_errors.append("Line item validation error processing '{}' worksheet. {} is not a valid line item type. Valid line item types are: {}".format(sheet, tmpstr, valid_types))

    return tmpstr


def excel_to_checklist(data):
    # we now support the Generic_Checklist.xlsx template, and suggest 20 sections and 80 line items per section
    MAX_SUPPORTED_CHECKLIST_LINES=80
    MAX_SUPPORTED_CHECKLIST_LINES+=6 # for the xls offset-lines
    checklist_data = {}
    checklist_data['checklist_payload'] = {}
    checklist_payload = checklist_data['checklist_payload']
    checklist_payload['metadata'] = {}
    checklist_payload['sections'] = []
    checklist_payload['section'] = {}
    meta_section = checklist_payload['metadata']
    checklist_data['checklist_errors'] = []
    checklist_errors = checklist_data['checklist_errors']

    #---------------------------------
    # Handle the 'Instructions' Sheet
    try:
        page = data['Instructions']
        # Define the metadata values 'checklist_name', 'checklist_version', 'report_type', 'checklist_type', 'default_accessories', 'accessories'
        for count in range(1, 30):
            if page['A{}'.format(count)].value and page['B{}'.format(count)].value:
                if page['A{}'.format(count)].value.lower() == "checklist name":
                    if page['B{}'.format(count+1)].value:
                        if page['B{}'.format(count)].value and page['B{}'.format(count+1)].value:
                            if 'observation' in page['B{}'.format(count+2)].value.lower():
                                meta_section['checklist_name'] = "{}".format(page['B{}'.format(count)].value)
                            else:
                                meta_section['checklist_name'] = "{} [{}]".format(page['B{}'.format(count)].value, page['B{}'.format(count+1)].value)
                        else:
                            meta_section['checklist_name'] = "{}".format(page['B{}'.format(count)].value)
                if page['A{}'.format(count)].value.lower() == "checklist version":
                    meta_section['checklist_version'] = page['B{}'.format(count)].value
                if page['A{}'.format(count)].value.lower() == "report type":
                    meta_section['report_type'] = page['B{}'.format(count)].value
                if page['A{}'.format(count)].value.lower() == "checklist type":
                    meta_section['checklist_type'] = page['B{}'.format(count)].value
                if 'accessories' in page['A{}'.format(count)].value.lower().split():
                    meta_section['default_accessories'] = validate_accessories(page['B{}'.format(count)].value, '0')
        # Validate the metadata section by keys
        if meta_section:
            valid_metadata = ['checklist_name', 'checklist_version', 'checklist_type', 'default_accessories']
            if not all(keys in meta_section for keys in valid_metadata):
                checklist_errors.append('Missing metadata. Metadata section requires {}, but currently only contains {}'.format(meta_section, valid_metadata))
    except Exception as e:
        logging.error("ERROR: {}".format(e))
        checklist_errors.append('The Instructions page has an issue: {}'.format(e))

    #---------------------------------
    # Handle the 'Section#' Sheets
    # go through sheets in workbook and grab all that start with "Section"
    sheets = [i.strip() for i in data.sheetnames if i.startswith("Section") == True]
    for sheet in sheets:
        sheet_slug = slugify(sheet, separator='_')
        # if they have data, add it to the checklist
        page = data[sheet]
        if page['B2'].value:
            checklist_payload['sections'].append(sheet_slug)
            checklist_payload['section'][sheet_slug] = {
                'ordered_items': [],
                'display_text': page['B2'].value.strip(),
                'line_items': {}
            }
            # set section_data to make iterating through dictionary more clear
            section_data = checklist_payload['section'][sheet_slug]
            # loop through the line items in this section, if the row is filled out go ahead and add it
            col_name_lookup={ 'item_name': 'B',
                              'type': 'C',
                              'choices': 'D',
                              'accessories_on': 'E',
                              'accessories_off': 'F',
                              'description': 'G',
                              'image_source': 'H'}

            line_item_count=5
            for count in range(6, MAX_SUPPORTED_CHECKLIST_LINES):
                section_item_count = 0
                # back to regular processing
                if not page["{}{}".format(col_name_lookup['type'], count)].value:
                    line_item_count+=1
                    if page["{}{}".format(col_name_lookup['item_name'], count)].value:
                        checklist_errors.append("The type column must have a value in column {} {}".format(col_name_lookup['item_name'], count).value)
                        logging.info("ERROR: type column does not have a value.")
                else:
                    # Validate the type
                    line_item_type = validate_line_item_type(str(page["{}{}".format(col_name_lookup['type'], count)].value), sheet)
                    # set the 'display_text' and the validation type
                    display_text=""
                    if page["{}{}".format(col_name_lookup['item_name'], count)].value:
                        tmpstr = "".join([x if ord(x) < 128 else '?' for x in page['%s%s'% (col_name_lookup['item_name'], count)].value])
                        display_text = tmpstr.strip()
                    else:
                        checklist_errors.append('Every line item must have a name. Check {} {} in {}'.format(col_name_lookup['item_name'], count, page))
                    #---------------------------------------------------------------------------
                    # Do these operations on every line item
                    #---------------------------------------------------------------------------
                    line_item_count+=1
                    display_text=""
                    if page['%s%s'% (col_name_lookup['item_name'],line_item_count)].value:
                        tmpstr = "".join([x if ord(x) < 128 else '?' for x in page['%s%s'% (col_name_lookup['item_name'],count)].value])
                        display_text = tmpstr.strip()
                        display_slug = slugify(display_text, separator='_')

                    if display_slug in section_data['line_items']:
                        line_item_name = '{}_{}_{}'.format(sheet, display_slug, section_item_count)
                        section_item_count += 1
                    else:
                        line_item_name = '{}_{}'.format(sheet, display_slug)

                    section_data['ordered_items'].append(line_item_name)
                    section_data['line_items'][line_item_name] = {
                                                                    'display_text': display_text,
                                                                    'notes': '',
                                                                    'photo': [],
                                                                    'validation': False,
                                                                    'validation_type': line_item_type,
                                                                    'results': {}
                                                                }
                    # set section_line_items to make iterating through the dictionary more clear
                    section_line_items = section_data['line_items'][line_item_name]
                    # look for description
                    if page['%s%s'% (col_name_lookup['description'],line_item_count)].value:
                        tmpstr = "".join([x if ord(x) < 128 else '?' for x in page['%s%s'% (col_name_lookup['description'],line_item_count)].value])
                        description_text = tmpstr.strip()
                        section_line_items['description_text'] = description_text

                    if page['%s%s'% (col_name_lookup['accessories_on'],line_item_count)].value and page['%s%s'% (col_name_lookup['accessories_off'],line_item_count)].value:
                        logging.error("ERROR: line item [{}:{}] it is a misconfiguration to define both 'accessories_on' and 'accessories_off' at the same time".format(line_item_count, display_text))
                        checklist_errors.append("Update Line item [{}:{}] it is a misconfiguration to define both 'accessories_on' and 'accessories_off' at the same time".format(line_item_count, display_text))

                    if page['%s%s'% (col_name_lookup['accessories_on'],line_item_count)].value:
                        section_line_items['accessories_on'] = validate_accessories(page['%s%s'% (col_name_lookup['accessories_on'],line_item_count)].value, sheet)

                    if page['%s%s'% (col_name_lookup['accessories_off'],line_item_count)].value:
                        section_line_items['accessories_off'] = validate_accessories(page['%s%s'% (col_name_lookup['accessories_off'],line_item_count)].value, sheet)

                    #------------------- 'text' -------------------- 'text date' ----------- 'text block'----------
                    if line_item_type == 'text' or line_item_type == 'text date' or line_item_type == 'text block':
                        # actually handled by default rules above
                        pass
                    #------------------- 'choice' -------------------- 'multiple choice' ---------------
                    elif line_item_type == 'choice' or line_item_type == 'multiple choice':
                        section_line_items['choices'] = {}
                        # look for commas, break string into an array
                        if page['%s%s'% (col_name_lookup['choices'],line_item_count)].value == None:
                            logging.error("ERROR: the 'choice' type col must have a valid 'choices' [{}:{}] [{}]".format(sheet, count, line_item_type))
                            checklist_errors.append('Invalid entry in choice column [{}:{}] [{}]'.format(sheet, count, line_item_type))
                        else:
                            tmp_array = page['%s%s'% (col_name_lookup['choices'],line_item_count)].value.split(',')
                            # ditch spaces,
                            options=[]
                            for item in tmp_array:
                                if len(item) > 34:
                                    logging.error("ERROR: the 'choice' [{}] length is greater than 33 characters' [{}:{}] [{}]".format(item.lstrip(), sheet, count, line_item_type))
                                    checklist_errors.append('The choice [{}] cannot be more than 33 characters long. [{}:{}] [{}]'.format(item.lstrip(), sheet, count, line_item_type))
                                options.append(item.lstrip())
                                # loop through the array, left-padding line_item_count to create the dictionary
                                for n in range(0, len(options)):
                                    if n < 10:
                                        str_n = '0%s' % n
                                    else:
                                        str_n = str(n)
                                    section_line_items['choices'][str_n] = options[n]

                    #------------------- 'image' -------------------
                    elif line_item_type == 'image':
                        if page['%s%s'% (col_name_lookup['image_source'],line_item_count)].value == None:
                            logging.error("ERROR: 'image' type must have a valid 'src_url' [{}:{}] {}".format(sheet, count, line_item_type))
                            checklist_errors.append('The image must have a src_url.[{}:{}] {}".format(sheet, count, line_item_type)')
                        # elif the url is not valid - return error
                        else:
                            section_line_items['image_src'] = page['%s%s'% (col_name_lookup['image_source'],line_item_count)].value.strip()

                    #------------------- 'signature' -------------------
                    elif line_item_type == 'signature':
                        pass
        else:
            logging.warn("WARNING: Found Sheet [" + sheet + "] with no 'Section Name' defined: skipping this worksheet")
            # this section doesn't have a name so assume its blank and move on to next iteration
            continue

    return checklist_data
