import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
emails = {}
non_standard_emails = {}
no_change_emails = {}
new_emails = {}

wb = load_workbook("/Users/juliaguimiot/Downloads/eon_users.xlsx")
sheet = wb["Non standard email changes"]
for value in sheet.iter_rows(min_row=1, values_only=True):
    current_email = value[2]
    email = {
        "first_name": value[0],
        "last_name": value[1],
        "current_email": current_email,
        "new_email": value[3],
    }
    non_standard_emails[current_email] = email

sheet = wb["No Change"]

for value in sheet.iter_rows(min_row=1, values_only=True):
    current_email = value[2]
    email = {
        "first_name": value[0],
        "last_name": value[1],
        "current_email": current_email,
    }
    no_change_emails[current_email] = email

sheet = wb["New Users"]
for value in sheet.iter_rows(min_row=1, values_only=True):
    current_email = value[2]
    email = {
        "first_name": value[0],
        "last_name": value[1],
        "email": current_email,
        "role": value[3],
    }
    new_emails[current_email] = email

emails["non_standard"] = non_standard_emails
emails["no_change"] = no_change_emails
emails["new_emails"] = new_emails

with open("RWE_Emails.json", "w") as writeJsonfile:
      json.dump(emails, writeJsonfile, indent=4,default=str)
