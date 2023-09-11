import os
import os.path
import sys
from openpyxl import Workbook, load_workbook
from io import BytesIO

file = '/Users/juliaguimiot/Desktop/standalone_checklist_export/templates/Generic_Checklist.xlsx'
wb = Workbook()
wb = load_workbook(BytesIO(file))
test_checklist = excel_to_checklist(wb)
