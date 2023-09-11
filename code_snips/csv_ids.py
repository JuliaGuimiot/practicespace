import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("Blade_Inspection_defects.xlsx")
ws = wb.get_sheet_by_name('Blade_Inspection_defects')
column = ws['N']
column_list = [column[x].value for x in range(len(column))]
column_set = set(column_list)
column_list = list(column_set)
column_list.remove('project_id')
column_list.remove(None)

with open('ids.txt', 'w') as f:
    for item in column_list:
        f.write('%s\n' % item)
        print "-" * 100
        print("Wrote to file {}".format(f))
        print "-" * 100
